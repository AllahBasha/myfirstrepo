import array
import openpyxl


# def read_excel_data():

file = "C:\\Users\\LENOVO-T470\\Desktop\\Testcase\\Book2.xlsx"
# sheet = "Sheet1"
work_book = openpyxl.load_workbook(file)
print(work_book)
sheet = work_book["Sheet1"]
print(sheet)
row_count = sheet.max_row
col_count = sheet.max_column
print(row_count, col_count)
# data = array.array('str',[])
rows = []
data = []
for r in range(2,row_count+1):
    for c in range(1,col_count+1):
        rows.append(sheet.cell(r,c).value)
    data.append(tuple(rows))
    rows.clear()
    print(data)
    # return data

