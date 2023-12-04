import openpyxl
from selenium import webdriver
import pytest
from POM.LoginPage import Login


def read_excel_data():

    file = "C:\\Users\\LENOVO-T470\\Desktop\\Testcase\\Book2.xlsx"
    # sheet = "Sheet1"
    work_book = openpyxl.load_workbook(file)
    print(work_book)
    sheet = work_book["Sheet1"]
    print(sheet)
    row_count = sheet.max_row
    col_count = sheet.max_column
    print(row_count, col_count)
    # data = array.array('str',[])
    rows = []
    data = []
    for r in range(2, row_count + 1):
        for c in range(1, col_count + 1):
            rows.append(sheet.cell(r, c).value)
        data.append(tuple(rows))
        print(data)
        rows.clear()
    return data


class Test_001_Login:
    # url_link = "https://admin-demo.nopcommerce.com/login?ReturnUrl=%2Fadmin%2F"
    # username = "admin@yourstore.com"
    # password = "admin"

    test_data = read_excel_data()
    print("---------")
    print(test_data)

    def test_homepage_title(self, setup):
        # self.driver = webdriver.Chrome()
        # self.driver.get(self.url_link)
        self.driver = setup
        actual_title = self.driver.title
        self.driver.close()
        print(actual_title)
        if actual_title == "Your store. Login":
            assert True
        else:
            assert False

    @pytest.mark.parametrize("username, password", test_data)
    def test_login(self, username, password, setup):
        # self.driver = webdriver.Chrome()
        # self.driver.get(self.url_link)
        self.driver = setup
        data = read_excel_data()
        self.lp = Login(self.driver)
        self.lp.set_user_name(username)
        self.lp.set_password(password)
        self.lp.click_login()
        self.lp.click_logout()
        self.driver.close()
